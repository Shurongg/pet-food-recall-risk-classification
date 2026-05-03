"""
Phase 1 / 1.5 / 1.6 / 2.5 — Data Collection (multi-source)

Collects raw recall / food-alert data from all available sources and saves
everything to data/raw/ without filtering or cleaning.
Scope filtering happens in Phase 2 (inspect_data.py).

Sources attempted (in order):
  ── US sources ──────────────────────────────────────────────────────────────
  1. FDA CVM XLSX  (primary, always run)
     https://www.fda.gov/animal-veterinary/safety-health/recalls-withdrawals
  2. Data.gov "FDA Pet Food Recalls" XLS  (legacy, attempted, documented if 404)
     https://catalog.data.gov/dataset/fda-pet-food-recalls
  3. openFDA Food Enforcement API  (broad multi-keyword, deduplicated)
     https://api.fda.gov/food/enforcement.json
  ── European sources (Phase 1.6) ────────────────────────────────────────────
  4. UK FSA Food Alerts  (full bulk JSON + CSV, open licence OGL v3)
     https://data.food.gov.uk/food-alerts/id.json
  5. EU RASFF  (documented — requires EU Login authentication, no public bulk API)
     https://webgate.ec.europa.eu/rasff-window/screen/search
  6. German BVL recalls  (documented — no structured public API found)
     https://www.bvl.bund.de
  7. French ANSES alerts  (documented — no structured public API found)
     https://www.anses.fr
  ── Canada / Oceania sources (Phase 2.5) ────────────────────────────────────
  8. Canada CFIA + Consumer Product Safety recalls JSON  (Open Government Licence)
     https://recalls-rappels.canada.ca
  9. Australia FSANZ Food Recalls  (documented — RSS only 10 recent items, no bulk export)
     https://www.foodstandards.gov.au/food-recalls
 10. New Zealand MPI Food Recalls  (documented — Incapsula bot-protection blocks access)
     https://www.mpi.govt.nz/food-safety-home/food-recalls-and-complaints/
 11. Australia ACCC Product Safety  (documented — no public JSON/CSV API found)
     https://www.productsafety.gov.au/recalls

Outputs written to --output (default: data/raw/):
  fda_animal_veterinary_recalls.xlsx        — raw CVM XLSX bytes
  fda_animal_veterinary_recalls.csv         — CVM data as CSV
  openfda_pet_food_enforcement_raw.csv      — openFDA deduplicated records
  uk_fsa_food_alerts_raw.csv               — all UK FSA food alerts (raw)
  canada_cfia_recalls_raw.csv              — Canada CFIA + Consumer Product Safety (Phase 2.5)
  source_inventory.csv                      — one row per source/query
  source_metadata.json                      — full provenance record
"""

import argparse
import csv
import json
import os
import sys
from datetime import datetime, timezone

import requests


# ---------------------------------------------------------------------------
# Source constants
# ---------------------------------------------------------------------------

# Source 1 — FDA CVM XLSX
FDA_XLSX_URL = (
    "https://www.fda.gov/animal-veterinary/safety-health/recalls-withdrawals"
    "/datatables-data/download?gdpid=63326547d8114d0001348f0d&_format=xlsx"
)
FDA_PAGE_URL = "https://www.fda.gov/animal-veterinary/safety-health/recalls-withdrawals"

# Source 2 — Data.gov legacy XLS (accessdata.fda.gov, may be 404)
DATAGOV_PAGE_URL = "https://catalog.data.gov/dataset/fda-pet-food-recalls"
DATAGOV_XLS_URL = (
    "https://www.accessdata.fda.gov/scripts/newpetfoodrecalls/"
    "PetFoodRecallProductsList2009.xls"
)

# Source 3 — openFDA Food Enforcement API
OPENFDA_URL = "https://api.fda.gov/food/enforcement.json"

# Single-term (unquoted) queries that may surface pet food records.
# Broad by design — scope filtering is done in Phase 2.
OPENFDA_PET_QUERIES = [
    "dog",
    "cat",
    "pet",
    "treat",
    "chew",
    "puppy",
    "kitten",
]

OPENFDA_PAGE_SIZE = 100
OPENFDA_MAX_PER_QUERY = 500   # safety cap per keyword

# Source 4 — UK FSA Food Alerts (Linked Data API, open licence OGL v3)
# Full bulk fetch in one call — 1,314 total records as of 2026-05-03.
UK_FSA_API_URL = "https://data.food.gov.uk/food-alerts/id.json"
UK_FSA_PAGE_URL = "https://data.food.gov.uk/food-alerts/id"
UK_FSA_LICENCE = "OGL v3 — http://www.nationalarchives.gov.uk/doc/open-government-licence/version/3/"
# Generous limit to fetch all records; actual total ~1,300. Cap at 10,000 for safety.
UK_FSA_LIMIT = 10_000

# Source 5 — EU RASFF (documented; requires EU Login)
RASFF_PORTAL_URL = "https://webgate.ec.europa.eu/rasff-window/screen/search"
RASFF_BACKEND_URL = "https://webgate.ec.europa.eu/rasff-window/backend"

# Source 6 — German BVL (documented; no public API)
BVL_PAGE_URL = "https://www.bvl.bund.de/DE/Arbeitsbereiche/01_Lebensmittel/04_Schnellwarnsysteme/bvl_schnellwarnungen_node.html"

# Source 7 — French ANSES (documented; no public API)
ANSES_PAGE_URL = "https://www.anses.fr/fr/content/les-alertes-alimentaires"

# Source 8 — Canada CFIA + Consumer Product Safety recalls JSON (Phase 2.5)
# Static JSON file updated daily; no authentication required.
# Licence: Open Government Licence - Canada v2.0
CANADA_JSON_URL = (
    "https://recalls-rappels.canada.ca/sites/default/files/opendata-donneesouvertes/"
    "HCRSAMOpenData.json"
)
CANADA_PAGE_URL = "https://recalls-rappels.canada.ca/en"
CANADA_LICENCE = "Open Government Licence - Canada v2.0 — https://open.canada.ca/en/open-government-licence-canada"
# Save records from food/consumer-product organizations; exclude vehicles, medical devices, drugs.
# CFIA = Canadian Food Inspection Agency (food, feed, plant safety)
# Consumer product safety = Health Canada branch handling non-food consumer recalls including pet food
CANADA_KEEP_ORGS = frozenset({"CFIA", "Consumer product safety"})

# Source 9 — Australia FSANZ (Phase 2.5, documented)
# RSS feed exists but contains only the 10 most-recent items.  No bulk export API found.
FSANZ_RSS_URL = "https://www.foodstandards.gov.au/food-recalls-rss.xml"
FSANZ_PAGE_URL = "https://www.foodstandards.gov.au/food-recalls"

# Source 10 — New Zealand MPI (Phase 2.5, documented)
# Site returns Incapsula bot-protection challenge for non-browser requests.
NZ_MPI_PAGE_URL = "https://www.mpi.govt.nz/food-safety-home/food-recalls-and-complaints/"
NZ_MPI_ACVM_URL = "https://www.mpi.govt.nz/agriculture/agricultural-compounds-vet-medicines/acvm-non-compliance-and-recalls/acvm-recalls"

# Source 11 — Australia ACCC Product Safety (Phase 2.5, documented)
# Drupal site; JSON:API disabled; no bulk CSV/JSON export found.
ACCC_PAGE_URL = "https://www.productsafety.gov.au/recalls"

REQUEST_TIMEOUT = 30  # seconds

# FDA bot-detection requires browser-like headers; plain requests gets blocked.
REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,"
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
        "*/*;q=0.8"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": FDA_PAGE_URL,
}

# Fields to use as deduplication key for openFDA records (in priority order)
OPENFDA_ID_FIELDS = ("event_id", "recall_number")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _now_utc() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _write_json(path: str, record: dict) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(record, f, indent=2, ensure_ascii=False)
    print(f"  Written → {path}")


def _unique_key(rec: dict) -> str:
    """Return the best available deduplication key for an openFDA record."""
    for field in OPENFDA_ID_FIELDS:
        val = rec.get(field)
        if val:
            return str(val)
    return f"{rec.get('product_description','')[:80]}|{rec.get('recall_initiation_date','')}"


# ---------------------------------------------------------------------------
# Source 1 — FDA CVM XLSX (primary)
# ---------------------------------------------------------------------------

def collect_fda_cvm(output_dir: str) -> tuple[dict, dict]:
    """Download the FDA CVM XLSX and return (metadata_dict, inventory_row)."""
    import openpyxl

    print(f"\n[1/7] FDA CVM XLSX\n  {FDA_XLSX_URL}")

    try:
        session = requests.Session()
        session.headers.update(REQUEST_HEADERS)
        resp = session.get(FDA_XLSX_URL, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        content_type = resp.headers.get("content-type", "")
        if "spreadsheetml" not in content_type and "octet-stream" not in content_type:
            raise ValueError(
                f"Unexpected content-type '{content_type}' — "
                "FDA may have returned an error page instead of the XLSX."
            )
    except (requests.exceptions.RequestException, ValueError) as exc:
        print(f"  ERROR: {exc}", file=sys.stderr)
        print(
            f"  Manual fallback: visit {FDA_PAGE_URL} and download the table as\n"
            f"  Excel, then place it in {output_dir} as fda_animal_veterinary_recalls.xlsx",
            file=sys.stderr,
        )
        sys.exit(1)

    timestamp = _now_utc()

    xlsx_path = os.path.join(output_dir, "fda_animal_veterinary_recalls.xlsx")
    with open(xlsx_path, "wb") as f:
        f.write(resp.content)
    print(f"  XLSX saved → {xlsx_path}  ({len(resp.content):,} bytes)")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        print("  ERROR: XLSX is empty.", file=sys.stderr)
        sys.exit(1)

    headers = [str(h) if h is not None else "" for h in all_rows[0]]
    data_rows = all_rows[1:]
    row_count = len(data_rows)

    csv_path = os.path.join(output_dir, "fda_animal_veterinary_recalls.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in data_rows:
            writer.writerow([("" if v is None else str(v)) for v in row])
    print(f"  CSV saved  → {csv_path}  ({row_count} data rows)")

    metadata = {
        "source_name": "FDA CVM — Animal & Veterinary Recalls / Withdrawals",
        "source_url": FDA_PAGE_URL,
        "download_url": FDA_XLSX_URL,
        "retrieval_timestamp": timestamp,
        "row_count": row_count,
        "column_names": headers,
        "collection_method": "Direct HTTP GET of published XLSX export (browser User-Agent required)",
        "saved_files": [os.path.basename(xlsx_path), os.path.basename(csv_path)],
    }

    inventory_row = {
        "source_name": "FDA CVM Animal & Veterinary Recalls XLSX",
        "source_url_or_query": FDA_PAGE_URL,
        "rows_found": row_count,
        "rows_saved": row_count,
        "notes": "Primary source. All animal/veterinary recall types included; scope filter in Phase 2.",
    }

    return metadata, inventory_row


# ---------------------------------------------------------------------------
# Source 2 — Data.gov legacy XLS
# ---------------------------------------------------------------------------

def collect_datagov_xls(output_dir: str) -> tuple[dict, list[dict]]:
    """Attempt to download the Data.gov / FDA legacy pet food XLS.

    Documents as unavailable if the file returns 404.
    """
    print(f"\n[2/7] Data.gov FDA Pet Food Recalls XLS (legacy)\n  {DATAGOV_XLS_URL}")

    timestamp = _now_utc()
    inventory_rows = []

    try:
        session = requests.Session()
        session.headers.update(REQUEST_HEADERS)
        resp = session.get(DATAGOV_XLS_URL, timeout=REQUEST_TIMEOUT)
        if resp.status_code == 404:
            raise requests.exceptions.HTTPError("404 Not Found")
        resp.raise_for_status()
        content_type = resp.headers.get("content-type", "")
        if "html" in content_type.lower() and len(resp.content) < 50_000:
            raise ValueError(f"Got HTML response instead of XLS (content-type: {content_type})")
    except (requests.exceptions.RequestException, ValueError) as exc:
        print(f"  UNAVAILABLE: {exc}")
        metadata = {
            "source_name": "Data.gov / FDA Pet Food Recalls XLS (2009, legacy)",
            "source_url": DATAGOV_PAGE_URL,
            "download_url": DATAGOV_XLS_URL,
            "retrieval_timestamp": timestamp,
            "status": "unavailable",
            "error": str(exc),
            "note": "The accessdata.fda.gov XLS link found on Data.gov returns 404 as of 2026.",
        }
        inventory_rows.append({
            "source_name": "Data.gov FDA Pet Food Recalls XLS (legacy 2009)",
            "source_url_or_query": DATAGOV_XLS_URL,
            "rows_found": 0,
            "rows_saved": 0,
            "notes": f"Unavailable: {exc}",
        })
        return metadata, inventory_rows

    xls_path = os.path.join(output_dir, "datagov_fda_pet_food_recalls_2009.xls")
    with open(xls_path, "wb") as f:
        f.write(resp.content)
    print(f"  XLS saved → {xls_path}  ({len(resp.content):,} bytes)")

    row_count = -1
    try:
        import xlrd  # type: ignore
        wb = xlrd.open_workbook(xls_path)
        row_count = wb.sheet_by_index(0).nrows - 1
    except Exception:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xls_path)
            row_count = wb.active.max_row - 1
        except Exception:
            pass

    metadata = {
        "source_name": "Data.gov / FDA Pet Food Recalls XLS (2009, legacy)",
        "source_url": DATAGOV_PAGE_URL,
        "download_url": DATAGOV_XLS_URL,
        "retrieval_timestamp": timestamp,
        "row_count": row_count,
        "collection_method": "Direct HTTP GET of legacy XLS from accessdata.fda.gov",
        "saved_files": [os.path.basename(xls_path)],
    }
    inventory_rows.append({
        "source_name": "Data.gov FDA Pet Food Recalls XLS (legacy 2009)",
        "source_url_or_query": DATAGOV_XLS_URL,
        "rows_found": row_count,
        "rows_saved": row_count,
        "notes": "Legacy dataset from accessdata.fda.gov via Data.gov listing.",
    })
    return metadata, inventory_rows


# ---------------------------------------------------------------------------
# Source 3 — openFDA Food Enforcement API (multi-keyword, deduplicated)
# ---------------------------------------------------------------------------

def collect_openfda_pet_food(output_dir: str) -> tuple[dict, list[dict]]:
    """Query openFDA Food Enforcement API with multiple pet-food keywords.

    Records are deduplicated across queries by event_id / recall_number.
    """
    print(f"\n[3/7] openFDA Food Enforcement API (multi-keyword)\n  {OPENFDA_URL}")

    timestamp = _now_utc()
    inventory_rows = []
    seen: dict[str, dict] = {}

    for term in OPENFDA_PET_QUERIES:
        search_param = f"product_description:{term}"
        skip = 0
        fetched_this_query = 0
        total_reported = None

        while True:
            try:
                resp = requests.get(
                    OPENFDA_URL,
                    params={"search": search_param, "limit": OPENFDA_PAGE_SIZE, "skip": skip},
                    timeout=REQUEST_TIMEOUT,
                )
                resp.raise_for_status()
                data = resp.json()
            except requests.exceptions.RequestException as exc:
                print(f"  WARNING: request failed for '{term}' at skip={skip} — {exc}",
                      file=sys.stderr)
                break

            if data.get("error"):
                total_reported = 0
                break

            if total_reported is None:
                total_reported = data.get("meta", {}).get("results", {}).get("total", 0) or 0

            batch = data.get("results", [])
            if not batch:
                break

            for rec in batch:
                key = _unique_key(rec)
                if key not in seen:
                    seen[key] = rec

            skip += len(batch)
            fetched_this_query += len(batch)

            if skip >= total_reported or skip >= OPENFDA_MAX_PER_QUERY:
                break

        for rec in seen.values():
            if "_query_term" not in rec:
                rec["_query_term"] = term

        print(f"  '{term:<12}'  API total={total_reported or 0:>4}  "
              f"fetched={fetched_this_query:>4}  "
              f"running_unique={len(seen):>4}")

        inventory_rows.append({
            "source_name": "openFDA Food Enforcement API",
            "source_url_or_query": f"{OPENFDA_URL}?search=product_description:{term}",
            "rows_found": total_reported or 0,
            "rows_saved": fetched_this_query,
            "notes": (
                "Broad keyword query. Deduplicated by event_id/recall_number across all "
                "queries. Final unique total in openfda_pet_food_enforcement_raw.csv."
            ),
        })

    all_records = list(seen.values())
    total_unique = len(all_records)

    if total_unique == 0:
        print("  No records returned by any openFDA query.")
        return {
            "source_name": "openFDA Food Enforcement API",
            "source_url": OPENFDA_URL,
            "retrieval_timestamp": timestamp,
            "row_count": 0,
            "note": "No records matched any pet-food keyword query.",
        }, inventory_rows

    all_keys: list[str] = []
    seen_keys: set[str] = set()
    for rec in all_records:
        for k in rec:
            if k != "_query_term" and k not in seen_keys:
                all_keys.append(k)
                seen_keys.add(k)
    all_keys.append("_query_term")

    csv_path = os.path.join(output_dir, "openfda_pet_food_enforcement_raw.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=all_keys, extrasaction="ignore")
        writer.writeheader()
        for rec in all_records:
            writer.writerow({k: rec.get(k, "") for k in all_keys})

    print(f"  Combined CSV → {csv_path}  ({total_unique} unique records)")

    metadata = {
        "source_name": "openFDA Food Enforcement API — multi-keyword pet food query",
        "source_url": OPENFDA_URL,
        "retrieval_timestamp": timestamp,
        "row_count": total_unique,
        "column_names": all_keys,
        "queries_used": [f"product_description:{t}" for t in OPENFDA_PET_QUERIES],
        "deduplication_key": list(OPENFDA_ID_FIELDS),
        "collection_method": (
            f"Paginated GET per keyword ({OPENFDA_PAGE_SIZE}/page, max "
            f"{OPENFDA_MAX_PER_QUERY}/term); deduplicated by event_id/recall_number"
        ),
        "saved_files": [os.path.basename(csv_path)],
        "note": (
            "Broad keyword queries capture human food records (e.g. 'hot dog buns'). "
            "Scope filtering to pet-food-only records is done in Phase 2."
        ),
    }
    return metadata, inventory_rows


# ---------------------------------------------------------------------------
# Source 4 — UK FSA Food Alerts (European, open API, OGL v3)
# ---------------------------------------------------------------------------

def collect_uk_fsa(output_dir: str) -> tuple[dict, dict]:
    """Download all UK Food Standards Agency food alert records.

    The FSA Linked Data API provides bulk JSON export without authentication.
    All records are saved raw; pet-food scope filtering is done in Phase 2.

    API: https://data.food.gov.uk/food-alerts/id.json
    Licence: OGL v3
    """
    api_url = f"{UK_FSA_API_URL}?_limit={UK_FSA_LIMIT}"
    print(f"\n[4/7] UK FSA Food Alerts (European)\n  {api_url}")

    timestamp = _now_utc()

    try:
        resp = requests.get(
            api_url,
            headers={"User-Agent": REQUEST_HEADERS["User-Agent"], "Accept": "application/json"},
            timeout=REQUEST_TIMEOUT,
        )
        resp.raise_for_status()
        data = resp.json()
    except (requests.exceptions.RequestException, ValueError) as exc:
        print(f"  ERROR: {exc}", file=sys.stderr)
        metadata = {
            "source_name": "UK Food Standards Agency — Food Alerts (Linked Data API)",
            "source_url": UK_FSA_PAGE_URL,
            "retrieval_timestamp": timestamp,
            "status": "error",
            "error": str(exc),
        }
        inv_row = {
            "source_name": "UK FSA Food Alerts (Linked Data API)",
            "source_url_or_query": api_url,
            "rows_found": 0,
            "rows_saved": 0,
            "notes": f"Error: {exc}",
        }
        return metadata, inv_row

    items = data.get("items", [])
    row_count = len(items)
    print(f"  Received {row_count} alert records")

    if row_count == 0:
        print("  WARNING: no records returned.", file=sys.stderr)
        metadata = {
            "source_name": "UK FSA Food Alerts",
            "source_url": UK_FSA_PAGE_URL,
            "retrieval_timestamp": timestamp,
            "row_count": 0,
            "note": "API returned 0 records.",
        }
        inv_row = {
            "source_name": "UK FSA Food Alerts (Linked Data API)",
            "source_url_or_query": api_url,
            "rows_found": 0,
            "rows_saved": 0,
            "notes": "API returned 0 records",
        }
        return metadata, inv_row

    # Flatten nested fields to a consistent CSV representation.
    # Nested lists/dicts are serialised as pipe-separated strings.
    def _flatten(item: dict) -> dict:
        flat: dict[str, str] = {}
        for k, v in item.items():
            if isinstance(v, list):
                flat[k] = " | ".join(str(x) for x in v)
            elif isinstance(v, dict):
                flat[k] = str(v)
            else:
                flat[k] = "" if v is None else str(v)
        return flat

    flat_items = [_flatten(item) for item in items]

    # Collect all column names seen (union)
    all_keys: list[str] = []
    seen_keys: set[str] = set()
    for item in flat_items:
        for k in item:
            if k not in seen_keys:
                all_keys.append(k)
                seen_keys.add(k)

    csv_path = os.path.join(output_dir, "uk_fsa_food_alerts_raw.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=all_keys, extrasaction="ignore")
        writer.writeheader()
        for row in flat_items:
            writer.writerow({k: row.get(k, "") for k in all_keys})

    print(f"  CSV saved  → {csv_path}  ({row_count} rows)")

    metadata = {
        "source_name": "UK Food Standards Agency — Food Alerts (Linked Data API)",
        "source_url": UK_FSA_PAGE_URL,
        "api_url": api_url,
        "licence": UK_FSA_LICENCE,
        "retrieval_timestamp": timestamp,
        "row_count": row_count,
        "column_names": all_keys,
        "collection_method": (
            f"Single HTTP GET to FSA Linked Data API with _limit={UK_FSA_LIMIT} "
            "(no authentication required). Nested fields serialised as pipe-separated strings."
        ),
        "saved_files": [os.path.basename(csv_path)],
        "note": (
            "Covers all FSA product alerts including human food, pet food, and feed. "
            "~37 records mention pet/dog/cat food in title or product name. "
            "Scope filtering applied in Phase 2."
        ),
    }

    inv_row = {
        "source_name": "UK FSA Food Alerts (Linked Data API)",
        "source_url_or_query": api_url,
        "rows_found": row_count,
        "rows_saved": row_count,
        "notes": (
            "Full bulk download of all FSA food alerts (human + pet + feed, unfiltered). "
            f"Licence: {UK_FSA_LICENCE}. Scope filtering in Phase 2."
        ),
    }

    return metadata, inv_row


# ---------------------------------------------------------------------------
# Sources 5–7 — EU RASFF, German BVL, French ANSES  (documented only)
# ---------------------------------------------------------------------------

def document_eu_inaccessible_sources() -> tuple[dict, list[dict]]:
    """Return metadata and inventory rows for EU sources that were attempted
    but found to have no publicly accessible bulk data API.

    No HTTP requests are made; these are documentary records only.
    """
    timestamp = _now_utc()

    sources = [
        {
            "key": "eu_rasff",
            "source_name": "EU RASFF (Rapid Alert System for Food and Feed)",
            "source_url": RASFF_PORTAL_URL,
            "retrieval_timestamp": timestamp,
            "status": "access_restricted",
            "access_finding": (
                "RASFF portal is an Angular SPA. Backend API endpoints "
                f"({RASFF_BACKEND_URL}/notification/search/consolidated, "
                f"{RASFF_BACKEND_URL}/notification/search/export) return the Angular "
                "shell HTML for unauthenticated requests — EU Login (iRASFF) is required. "
                "No public bulk-download endpoint was found. Annual report PDFs are "
                "published but contain no machine-readable data."
            ),
            "access_method_required": "EU Login (iRASFF system)",
            "note": (
                "RASFF is the primary EU food safety alert system covering all 27 EU member "
                "states plus EEA/UK/Switzerland. It would be the best EU source if authenticated "
                "access were available. Future work could request data access via the EC."
            ),
            "inv_row": {
                "source_name": "EU RASFF — Rapid Alert System for Food and Feed",
                "source_url_or_query": RASFF_PORTAL_URL,
                "rows_found": 0,
                "rows_saved": 0,
                "notes": (
                    "Access restricted: requires EU Login / iRASFF authentication. "
                    "Backend API returns Angular SPA shell for unauthenticated requests. "
                    "No public bulk-download endpoint found."
                ),
            },
        },
        {
            "key": "german_bvl",
            "source_name": "German BVL (Bundesamt für Verbraucherschutz und Lebensmittelsicherheit)",
            "source_url": BVL_PAGE_URL,
            "retrieval_timestamp": timestamp,
            "status": "no_public_api",
            "access_finding": (
                "BVL publishes food safety rapid alerts on its website but no structured "
                "public API or bulk CSV/JSON download was found. The website returns "
                "session-based redirects (HTTP 303) for direct file access attempts."
            ),
            "note": "Alert summaries are available as HTML pages only.",
            "inv_row": {
                "source_name": "German BVL — Bundesamt für Verbraucherschutz",
                "source_url_or_query": BVL_PAGE_URL,
                "rows_found": 0,
                "rows_saved": 0,
                "notes": (
                    "No structured public API found. Website returns session-based "
                    "redirects for file access. Alerts are HTML-only."
                ),
            },
        },
        {
            "key": "french_anses",
            "source_name": "French ANSES (Agence nationale de sécurité sanitaire)",
            "source_url": ANSES_PAGE_URL,
            "retrieval_timestamp": timestamp,
            "status": "no_public_api",
            "access_finding": (
                "ANSES food alerts page returned 404 for attempted API endpoints. "
                "No structured public bulk-download API was found."
            ),
            "note": "Alerts are published as HTML/PDF press releases only.",
            "inv_row": {
                "source_name": "French ANSES — Agence nationale de sécurité sanitaire",
                "source_url_or_query": ANSES_PAGE_URL,
                "rows_found": 0,
                "rows_saved": 0,
                "notes": (
                    "No structured public API found. "
                    "Alerts published as HTML/PDF press releases only."
                ),
            },
        },
    ]

    metadata: dict[str, dict] = {}
    inventory_rows: list[dict] = []

    for src in sources:
        key = src.pop("key")
        inv_row = src.pop("inv_row")
        metadata[key] = src
        inventory_rows.append(inv_row)
        status = src.get("status", "")
        label = "ACCESS RESTRICTED" if status == "access_restricted" else "NO PUBLIC API"
        print(f"  {label}: {src['source_name'][:60]}")

    return metadata, inventory_rows


# ---------------------------------------------------------------------------
# Source 8 — Canada CFIA + Consumer Product Safety recalls JSON (Phase 2.5)
# ---------------------------------------------------------------------------

def collect_canada_cfia(output_dir: str) -> tuple[dict, dict]:
    """Download Canada Health Canada / CFIA open recalls JSON.

    The JSON file is updated daily and contains all Canadian government recalls
    across many agencies.  We save only the CFIA (food/feed/plant) and Consumer
    product safety organizations (~10 k records); vehicles, drugs, and medical
    devices are excluded.  Scope filtering to pet-food-only records is done in
    Phase 2.

    Licence: Open Government Licence - Canada v2.0
    """
    print(f"\n[8/11] Canada CFIA + Consumer Product Safety recalls JSON\n  {CANADA_JSON_URL}")

    timestamp = _now_utc()

    try:
        resp = requests.get(
            CANADA_JSON_URL,
            headers={"User-Agent": REQUEST_HEADERS["User-Agent"], "Accept": "application/json"},
            timeout=REQUEST_TIMEOUT,
        )
        resp.raise_for_status()
        all_records = resp.json()
    except (requests.exceptions.RequestException, ValueError) as exc:
        print(f"  ERROR: {exc}", file=sys.stderr)
        metadata = {
            "source_name": "Canada Health Canada / CFIA Recalls and Safety Alerts JSON",
            "source_url": CANADA_PAGE_URL,
            "download_url": CANADA_JSON_URL,
            "retrieval_timestamp": timestamp,
            "status": "error",
            "error": str(exc),
        }
        inv_row = {
            "source_name": "Canada CFIA + Consumer Product Safety recalls (JSON)",
            "source_url_or_query": CANADA_JSON_URL,
            "rows_found": 0,
            "rows_saved": 0,
            "notes": f"Error: {exc}",
        }
        return metadata, inv_row

    if not isinstance(all_records, list):
        print("  ERROR: unexpected JSON structure (expected list)", file=sys.stderr)
        metadata = {
            "source_name": "Canada Health Canada / CFIA Recalls and Safety Alerts JSON",
            "source_url": CANADA_PAGE_URL,
            "download_url": CANADA_JSON_URL,
            "retrieval_timestamp": timestamp,
            "status": "error",
            "error": "Unexpected JSON structure — expected a list of recall records.",
        }
        inv_row = {
            "source_name": "Canada CFIA + Consumer Product Safety recalls (JSON)",
            "source_url_or_query": CANADA_JSON_URL,
            "rows_found": 0,
            "rows_saved": 0,
            "notes": "Unexpected JSON structure",
        }
        return metadata, inv_row

    total_in_json = len(all_records)
    print(f"  Downloaded {total_in_json:,} total records from JSON")

    # Keep only food / consumer-product organizations; drop vehicles, drugs, medical devices.
    subset = [
        r for r in all_records
        if (r.get("Organization") or "") in CANADA_KEEP_ORGS
    ]
    row_count = len(subset)
    print(f"  Filtered to {row_count:,} food/consumer-safety records "
          f"(orgs: {', '.join(sorted(CANADA_KEEP_ORGS))})")

    if row_count == 0:
        print("  WARNING: no records after organization filter.", file=sys.stderr)
        metadata = {
            "source_name": "Canada Health Canada / CFIA Recalls and Safety Alerts JSON",
            "source_url": CANADA_PAGE_URL,
            "download_url": CANADA_JSON_URL,
            "licence": CANADA_LICENCE,
            "retrieval_timestamp": timestamp,
            "row_count": 0,
            "note": "No records matched organization filter.",
        }
        inv_row = {
            "source_name": "Canada CFIA + Consumer Product Safety recalls (JSON)",
            "source_url_or_query": CANADA_JSON_URL,
            "rows_found": total_in_json,
            "rows_saved": 0,
            "notes": "No records after organization filter",
        }
        return metadata, inv_row

    keys = list(subset[0].keys())

    csv_path = os.path.join(output_dir, "canada_cfia_recalls_raw.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
        writer.writeheader()
        for row in subset:
            writer.writerow({k: (row.get(k) or "") for k in keys})

    print(f"  CSV saved  → {csv_path}  ({row_count:,} rows)")

    metadata = {
        "source_name": "Canada Health Canada / CFIA Recalls and Safety Alerts JSON",
        "source_url": CANADA_PAGE_URL,
        "download_url": CANADA_JSON_URL,
        "licence": CANADA_LICENCE,
        "retrieval_timestamp": timestamp,
        "total_records_in_json": total_in_json,
        "row_count": row_count,
        "column_names": keys,
        "organizations_included": sorted(CANADA_KEEP_ORGS),
        "organizations_excluded": "TC (transport), Medical devices, Drugs and health products, "
                                  "Marketed health products, Controlled substances and cannabis, HC",
        "collection_method": (
            f"Single HTTP GET of static daily-updated JSON file; "
            f"filtered to Organization in {{{', '.join(sorted(CANADA_KEEP_ORGS))}}}; "
            "no authentication required."
        ),
        "saved_files": [os.path.basename(csv_path)],
        "note": (
            "CFIA = Canadian Food Inspection Agency (food, feed, plant safety). "
            "'Consumer product safety' = Health Canada branch handling non-food consumer "
            "recalls including pet food.  Scope filtering to pet-food-only records in Phase 2."
        ),
    }

    inv_row = {
        "source_name": "Canada CFIA + Consumer Product Safety recalls (JSON)",
        "source_url_or_query": CANADA_JSON_URL,
        "rows_found": total_in_json,
        "rows_saved": row_count,
        "notes": (
            f"Total JSON records: {total_in_json:,}; saved {row_count:,} "
            f"(CFIA + Consumer product safety). Licence: {CANADA_LICENCE}. "
            "Phase 2 filters to pet-food scope."
        ),
    }

    return metadata, inv_row


# ---------------------------------------------------------------------------
# Sources 9–11 — Australia FSANZ, New Zealand MPI, Australia ACCC (documented)
# ---------------------------------------------------------------------------

def document_oceania_sources() -> tuple[dict, list[dict]]:
    """Return metadata and inventory rows for Oceania sources that were attempted
    but found to have no accessible bulk data API.

    No HTTP requests are made; these are documentary records only.

    Finding summary:
      FSANZ (AU): RSS feed confirmed at /food-recalls-rss.xml but contains only the
        10 most-recent items — insufficient for historical analysis.  Drupal JSON:API
        returns 404; full listing page shows only 3 items.  No bulk CSV/JSON export found.
      NZ MPI: All page requests return Incapsula bot-protection challenge (HTML iframe
        with NOINDEX/NOFOLLOW meta).  No structured data accessible without a real browser.
      ACCC (AU): Drupal site with JSON:API disabled (HTTP 404).  Search page returns
        HTML only.  No public JSON/CSV export found.
    """
    timestamp = _now_utc()
    print(f"\n[9–11/11] Oceania sources (documented)")

    sources = [
        {
            "key": "australia_fsanz",
            "source_name": "Australia FSANZ — Food Standards Australia New Zealand Food Recalls",
            "source_url": FSANZ_PAGE_URL,
            "retrieval_timestamp": timestamp,
            "status": "insufficient_data",
            "access_finding": (
                f"RSS feed ({FSANZ_RSS_URL}) confirmed accessible but contains only the "
                "10 most-recent recalls — insufficient for historical analysis.  "
                "Drupal JSON:API endpoint returns HTTP 404 (disabled).  "
                "Full listing page shows the same 3 most-recent items.  "
                "No bulk CSV/JSON export endpoint found."
            ),
            "note": (
                "FSANZ oversees food safety standards for both Australia and New Zealand.  "
                "A historical bulk export would be valuable but is not publicly available. "
                "RSS feed URL documented for future manual checking."
            ),
            "rss_url": FSANZ_RSS_URL,
            "inv_row": {
                "source_name": "Australia FSANZ Food Recalls",
                "source_url_or_query": FSANZ_PAGE_URL,
                "rows_found": 0,
                "rows_saved": 0,
                "notes": (
                    "Insufficient data: RSS feed confirmed but contains only 10 most-recent items.  "
                    "No bulk export or JSON API found."
                ),
            },
        },
        {
            "key": "new_zealand_mpi",
            "source_name": "New Zealand MPI — Ministry for Primary Industries Food Recalls",
            "source_url": NZ_MPI_PAGE_URL,
            "acvm_url": NZ_MPI_ACVM_URL,
            "retrieval_timestamp": timestamp,
            "status": "bot_protected",
            "access_finding": (
                "All HTTP requests to mpi.govt.nz return an Incapsula bot-protection challenge "
                "(HTML iframe with NOINDEX/NOFOLLOW meta, incident ID in response).  "
                "The ACVM recalls page (agricultural compounds, vet medicines, pet food, "
                "animal feed) is similarly inaccessible without a full browser session."
            ),
            "note": (
                "NZ MPI / ACVM would be a good source for pet food and animal feed recalls "
                "in New Zealand.  Data can only be accessed via a real browser session.  "
                "Future work could use a headless browser or manual export."
            ),
            "inv_row": {
                "source_name": "New Zealand MPI Food Recalls / ACVM",
                "source_url_or_query": NZ_MPI_PAGE_URL,
                "rows_found": 0,
                "rows_saved": 0,
                "notes": (
                    "Inaccessible: Incapsula bot-protection returns challenge page for "
                    "non-browser HTTP requests."
                ),
            },
        },
        {
            "key": "australia_accc",
            "source_name": "Australia ACCC — Product Safety Recalls",
            "source_url": ACCC_PAGE_URL,
            "retrieval_timestamp": timestamp,
            "status": "no_public_api",
            "access_finding": (
                f"ACCC Product Safety ({ACCC_PAGE_URL}) is a Drupal site.  "
                "JSON:API endpoint returns HTTP 404 (disabled).  "
                "Search page returns HTML only with no embedded structured data.  "
                "No public JSON/CSV bulk export endpoint found."
            ),
            "note": "Recalls are browsable via the web search interface but not programmatically exportable.",
            "inv_row": {
                "source_name": "Australia ACCC Product Safety Recalls",
                "source_url_or_query": ACCC_PAGE_URL,
                "rows_found": 0,
                "rows_saved": 0,
                "notes": "No public API: Drupal JSON:API disabled; no bulk export found.",
            },
        },
    ]

    metadata: dict[str, dict] = {}
    inventory_rows: list[dict] = []

    for src in sources:
        key = src.pop("key")
        inv_row = src.pop("inv_row")
        metadata[key] = src
        inventory_rows.append(inv_row)
        status_label = {
            "insufficient_data": "INSUFFICIENT DATA",
            "bot_protected":     "BOT PROTECTED",
            "no_public_api":     "NO PUBLIC API",
        }.get(src.get("status", ""), src.get("status", "").upper())
        print(f"  {status_label}: {src['source_name'][:60]}")

    return metadata, inventory_rows


# ---------------------------------------------------------------------------
# Source inventory writer
# ---------------------------------------------------------------------------

def write_source_inventory(output_dir: str, inventory_rows: list[dict]) -> str:
    """Merge new inventory rows with any existing source_inventory.csv.

    New rows for sources already listed (matched by source_name) replace the old
    row.  Sources in the existing file that were not re-run are preserved as-is.
    """
    fieldnames = ["source_name", "source_url_or_query", "rows_found", "rows_saved", "notes"]
    path = os.path.join(output_dir, "source_inventory.csv")

    existing: dict[str, dict] = {}
    if os.path.exists(path):
        try:
            with open(path, newline="", encoding="utf-8") as f:
                for row in csv.DictReader(f):
                    existing[row["source_name"]] = row
        except Exception:
            pass

    for row in inventory_rows:
        existing[row["source_name"]] = row

    merged = list(existing.values())
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(merged)
    print(f"\n  Inventory   → {path}  ({len(merged)} rows)")
    return path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Phase 1/1.5/1.6/2.5: Download raw pet food recall data from multiple sources.\n\n"
            "US sources (1–3):\n"
            "  1. FDA CVM Animal & Veterinary Recalls XLSX\n"
            "  2. Data.gov FDA Pet Food Recalls XLS (legacy; documented if unavailable)\n"
            "  3. openFDA Food Enforcement API (multi-keyword, deduplicated)\n\n"
            "European sources (4–7):\n"
            "  4. UK FSA Food Alerts (bulk JSON, OGL v3)\n"
            "  5. EU RASFF  — documented (requires EU Login)\n"
            "  6. German BVL — documented (no public API)\n"
            "  7. French ANSES — documented (no public API)\n\n"
            "Canada / Oceania sources (8–11):\n"
            "  8. Canada CFIA + Consumer Product Safety recalls JSON (Open Govt Licence)\n"
            "  9. Australia FSANZ — documented (RSS only 10 items, no bulk export)\n"
            " 10. New Zealand MPI — documented (Incapsula bot-protection)\n"
            " 11. Australia ACCC — documented (no public JSON/CSV API)\n\n"
            "Outputs in --output:\n"
            "  fda_animal_veterinary_recalls.xlsx/.csv\n"
            "  openfda_pet_food_enforcement_raw.csv\n"
            "  uk_fsa_food_alerts_raw.csv\n"
            "  canada_cfia_recalls_raw.csv\n"
            "  source_inventory.csv\n"
            "  source_metadata.json"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--output",
        default="data/raw/",
        help="Directory to save raw files (default: data/raw/)",
    )
    parser.add_argument(
        "--skip-fda-cvm",
        action="store_true",
        help="Skip the FDA CVM XLSX download",
    )
    parser.add_argument(
        "--skip-datagov",
        action="store_true",
        help="Skip the Data.gov legacy XLS attempt",
    )
    parser.add_argument(
        "--skip-openfda",
        action="store_true",
        help="Skip the openFDA Food Enforcement API queries",
    )
    parser.add_argument(
        "--skip-eu",
        action="store_true",
        help="Skip all European sources (UK FSA + EU documentation)",
    )
    parser.add_argument(
        "--skip-canada",
        action="store_true",
        help="Skip the Canada CFIA + Consumer Product Safety JSON download",
    )
    parser.add_argument(
        "--skip-oceania",
        action="store_true",
        help="Skip Oceania source documentation (FSANZ, NZ MPI, ACCC)",
    )
    # Backwards compatibility
    parser.add_argument(
        "--use-openfda-fallback",
        action="store_true",
        help=argparse.SUPPRESS,
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    os.makedirs(args.output, exist_ok=True)

    all_metadata: dict[str, dict] = {}
    all_inventory: list[dict] = []

    # --- US Source 1: FDA CVM ---
    if not args.skip_fda_cvm:
        cvm_meta, cvm_inv = collect_fda_cvm(args.output)
        all_metadata["fda_cvm"] = cvm_meta
        all_inventory.append(cvm_inv)
    else:
        print("\n[1/7] FDA CVM XLSX — skipped")

    # --- US Source 2: Data.gov legacy XLS ---
    if not args.skip_datagov:
        dg_meta, dg_inv_rows = collect_datagov_xls(args.output)
        all_metadata["datagov_legacy"] = dg_meta
        all_inventory.extend(dg_inv_rows)
    else:
        print("\n[2/7] Data.gov XLS — skipped")

    # --- US Source 3: openFDA ---
    if not args.skip_openfda:
        of_meta, of_inv_rows = collect_openfda_pet_food(args.output)
        all_metadata["openfda_enforcement"] = of_meta
        all_inventory.extend(of_inv_rows)
    else:
        print("\n[3/7] openFDA — skipped")

    # --- European Sources 4–7 ---
    if not args.skip_eu:
        # Source 4: UK FSA (data)
        fsa_meta, fsa_inv = collect_uk_fsa(args.output)
        all_metadata["uk_fsa"] = fsa_meta
        all_inventory.append(fsa_inv)

        # Sources 5–7: documented only
        print(f"\n[5–7/11] EU sources (documented)")
        eu_doc_meta, eu_doc_inv = document_eu_inaccessible_sources()
        all_metadata.update(eu_doc_meta)
        all_inventory.extend(eu_doc_inv)
    else:
        print("\n[4–7/11] European sources — skipped")

    # --- Canada source 8 ---
    if not args.skip_canada:
        ca_meta, ca_inv = collect_canada_cfia(args.output)
        all_metadata["canada_cfia"] = ca_meta
        all_inventory.append(ca_inv)
    else:
        print("\n[8/11] Canada CFIA — skipped")

    # --- Oceania sources 9–11 (documented only) ---
    if not args.skip_oceania:
        oc_meta, oc_inv = document_oceania_sources()
        all_metadata.update(oc_meta)
        all_inventory.extend(oc_inv)
    else:
        print("\n[9–11/11] Oceania sources — skipped")

    # --- Write outputs (merge with existing files so skipped sources are preserved) ---
    inv_path = write_source_inventory(args.output, all_inventory)
    meta_path = os.path.join(args.output, "source_metadata.json")
    # Merge: existing keys not touched by this run are preserved
    merged_meta: dict = {}
    if os.path.exists(meta_path):
        try:
            with open(meta_path, encoding="utf-8") as _f:
                merged_meta = json.load(_f)
        except Exception:
            pass
    merged_meta.update(all_metadata)
    _write_json(meta_path, merged_meta)

    # --- Terminal summary (use merged_meta so skipped sources still show correct counts) ---
    cvm_rows   = merged_meta.get("fda_cvm", {}).get("row_count", 0) or 0
    of_rows    = merged_meta.get("openfda_enforcement", {}).get("row_count", 0) or 0
    dg_status  = merged_meta.get("datagov_legacy", {}).get("status", "ok")
    dg_rows    = merged_meta.get("datagov_legacy", {}).get("row_count", 0) or 0
    fsa_rows   = merged_meta.get("uk_fsa", {}).get("row_count", 0) or 0
    fsa_status = merged_meta.get("uk_fsa", {}).get("status", "ok")
    ca_rows    = merged_meta.get("canada_cfia", {}).get("row_count", 0) or 0
    ca_status  = merged_meta.get("canada_cfia", {}).get("status", "ok")

    total_raw = (
        cvm_rows
        + of_rows
        + (dg_rows if dg_rows > 0 else 0)
        + (fsa_rows if fsa_status != "error" else 0)
        + (ca_rows if ca_status != "error" else 0)
    )

    print("\n" + "=" * 72)
    print("COLLECTION SUMMARY  (cumulative — skipped sources read from saved metadata)")
    print("=" * 72)
    print(f"  [1]  FDA CVM XLSX                {cvm_rows:>6} rows  → fda_animal_veterinary_recalls.csv")
    if dg_status == "unavailable":
        print(f"  [2]  Data.gov XLS           UNAVAILABLE  — documented in inventory")
    else:
        print(f"  [2]  Data.gov XLS                {dg_rows:>6} rows")
    print(f"  [3]  openFDA enforcement         {of_rows:>6} rows  → openfda_pet_food_enforcement_raw.csv")
    if fsa_status == "error":
        print(f"  [4]  UK FSA Food Alerts          ERROR  — documented in inventory")
    else:
        print(f"  [4]  UK FSA Food Alerts          {fsa_rows:>6} rows  → uk_fsa_food_alerts_raw.csv")
    print(f"  [5]  EU RASFF               ACCESS RESTRICTED — documented in inventory")
    print(f"  [6]  German BVL             NO PUBLIC API     — documented in inventory")
    print(f"  [7]  French ANSES           NO PUBLIC API     — documented in inventory")
    if ca_status == "error":
        print(f"  [8]  Canada CFIA               ERROR  — documented in inventory")
    else:
        print(f"  [8]  Canada CFIA               {ca_rows:>6} rows  → canada_cfia_recalls_raw.csv")
    print(f"  [9]  Australia FSANZ         INSUFFICIENT DATA — documented in inventory")
    print(f"  [10] New Zealand MPI         BOT PROTECTED     — documented in inventory")
    print(f"  [11] Australia ACCC          NO PUBLIC API     — documented in inventory")
    print(f"  {'─' * 60}")
    print(f"  Total raw rows saved            {total_raw:>6}  (pre-scope-filter)")
    print(f"  Source inventory        → {inv_path}")
    print(f"  Full metadata           → {meta_path}")
    print("=" * 72)
    print("  Phase 2 (inspect_data.py) will apply scope filtering.")
    print("=" * 72)


if __name__ == "__main__":
    main()
